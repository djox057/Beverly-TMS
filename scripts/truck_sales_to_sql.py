#!/usr/bin/env python3
"""Convert a Truck Sales spreadsheet into SQL UPDATE statements.

Usage:
    python scripts/truck_sales_to_sql.py "Copy of Truck Sales .xlsx" > backfill.sql

Options:
    --sheet NAME    Pick a specific worksheet (default: first sheet)
    --no-drivers    Skip driver UPDATEs
    --no-trucks     Skip truck UPDATEs

Expected sheet columns (in order):
    Truck# | Make | Model | Transmission | Year | Miles | Engine |
    APU/Webasto | Inverter | Fridge | Driver Name | Price (week) |
    Terms | Insurance (week) | Notes
"""
from __future__ import annotations

import argparse
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required. Install with: python -m pip install openpyxl")


def norm_truck(v):
    if v is None:
        return None
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else str(v)
    return str(v).strip() or None


def clean_text(v):
    if v is None:
        return None
    s = re.sub(r"\s+", " ", str(v).strip())
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s.upper() or None


def yn(v):
    if v is None:
        return False
    s = str(v).strip().upper()
    return s not in ("", "NO", "N", "/", "NONE")


def transmission(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    return "Automatic" if "auto" in s.lower() else s


def parse_int(v):
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def parse_price(v):
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "/"):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_terms_weeks(v):
    """Parse 'Ny Mm' (e.g. '3y 6m', '4Y', '2y 6m') into a week count."""
    if v is None:
        return None
    s = str(v).strip().lower()
    if s in ("", "/"):
        return None
    y = re.search(r"(\d+)\s*y", s)
    m = re.search(r"(\d+)\s*m", s)
    years = int(y.group(1)) if y else 0
    months = int(m.group(1)) if m else 0
    if years == 0 and months == 0:
        try:
            years = int(float(s))
        except (TypeError, ValueError):
            return None
    weeks = round(years * 52 + months * 52 / 12)
    return weeks if weeks > 0 else None


def sql_str(x):
    if x is None:
        return "NULL"
    return "'" + str(x).replace("'", "''") + "'"


def sql_num(x):
    return "NULL" if x is None else str(x)


def sql_bool(x):
    return "true" if x else "false"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--sheet", default=None)
    ap.add_argument("--no-trucks", action="store_true")
    ap.add_argument("--no-drivers", action="store_true")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]

    out = []
    out.append(f"-- Generated from {args.xlsx} / sheet: {ws.title}")
    out.append("BEGIN;")

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or len(row) < 13:
            continue
        tn = norm_truck(row[0])
        make_raw = row[1]
        if not tn or not make_raw:
            continue
        if tn.upper() in ("ONLY NEW UNITS",):
            continue

        make = clean_text(row[1])
        model = clean_text(row[2])
        tr = transmission(row[3])
        yr = parse_int(row[4])
        mi = parse_int(row[5])
        eng = clean_text(row[6])
        apu = yn(row[7])
        inv = yn(row[8])
        frg = yn(row[9])

        if not args.no_trucks:
            out.append(
                "UPDATE trucks SET "
                f"make={sql_str(make)}, model={sql_str(model)}, "
                f"transmission={sql_str(tr)}, year={sql_num(yr)}, "
                f"miles={sql_num(mi)}, engine={sql_str(eng)}, "
                f"has_apu_webasto={sql_bool(apu)}, "
                f"has_inverter={sql_bool(inv)}, "
                f"has_fridge={sql_bool(frg)} "
                f"WHERE truck_number={sql_str(tn)};"
            )

        if not args.no_drivers:
            price = parse_price(row[11])
            weeks = parse_terms_weeks(row[12])
            sets = []
            if price is not None:
                sets.append(f"weekly_payment={price}")
            if weeks is not None:
                sets.append(f"weeks_count={weeks}")
            if sets:
                out.append(
                    f"UPDATE drivers SET {', '.join(sets)} "
                    f"WHERE id = (SELECT driver1_id FROM trucks "
                    f"WHERE truck_number={sql_str(tn)}) "
                    f"AND id IS NOT NULL;"
                )

    out.append("COMMIT;")
    print("\n".join(out))


if __name__ == "__main__":
    main()